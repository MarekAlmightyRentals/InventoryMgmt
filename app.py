import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================
# CONFIG
# =====================
VENDOR_LIST = ["Crader Dist. (STIHL)"]
EXPECTED_FILES = ['activity', 'history', 'list']

st.set_page_config(
    page_title="\ud83d\udce6 Almighty Rentals | Inventory Optimizer",
    layout="wide"
)

st.markdown("<h1 style='text-align: center;'>\ud83d\udce6 Almighty Rentals Inventory Optimizer</h1>", unsafe_allow_html=True)
st.markdown("---")

# =====================
# HELPERS
# =====================
def normalize_columns(df, rename_map):
    df.columns = df.columns.str.strip().str.lower()
    for old, new in rename_map.items():
        if old in df.columns:
            df.rename(columns={old: new}, inplace=True)
    return df

def assign_abc(p):
    if p <= 0.2:
        return 'A'
    elif p <= 0.5:
        return 'B'
    return 'C'

def apply_excel_formatting(ws, col_idx):
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    grey = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    qty_cols = ['qty_sold', 'max_qty', 'min_qty', 're_order_point', 're_order_qty']

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sku = row[col_idx['sku'] - 1].value
        qty_sold_val = row[col_idx['qty_sold'] - 1].value

        # Highlight qty fields
        for qc in qty_cols:
            if qc in col_idx:
                row[col_idx[qc] - 1].fill = yellow

        # Row coloring
        fill = grey
        if sku:
            if sku.startswith("NS") and qty_sold_val and qty_sold_val > 0:
                fill = red
            elif sku.startswith("S-") and qty_sold_val and qty_sold_val > 0:
                fill = green

        for cell in row:
            cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws['B2']

# =====================
# UI SECTIONS
# =====================

def vendor_selector():
    vendor = st.selectbox("Select Vendor", VENDOR_LIST, index=0)
    st.markdown(f"\ud83d\udcdd Selected vendor: **{vendor}**")
    return vendor

def file_uploader_section():
    st.subheader("\ud83d\udcc2 Upload All Excel Files (Drag & Drop Supported):")
    uploaded_files = st.file_uploader("Upload all 3 Excel files here", type="xlsx", accept_multiple_files=True)
    st.caption("\u2705 Expected files:\n- Merchandise Activity.xlsx\n- Merchandise History.xlsx\n- Merchandise List.xlsx")
    return uploaded_files

# =====================
# APP LOGIC
# =====================
vendor = vendor_selector()
st.markdown("---")

st.subheader("\u2699\ufe0f Custom Stocking Parameters")
max_factor = st.slider("Max stock coverage (% of annual sales)", 10, 100, 50) / 100
min_factor = st.slider("Min stock threshold (% of max)", 5, 50, 25) / 100
reorder_point_factor = st.slider("Reorder point (% of max)", 5, 50, 25) / 100
st.markdown("---")

uploaded_files = file_uploader_section()
activity_file = history_file = list_file = None

if uploaded_files:
    file_map = {key: None for key in EXPECTED_FILES}
    for file in uploaded_files:
        fname = file.name.lower()
        for key in file_map:
            if key in fname:
                file_map[key] = file

    activity_file, history_file, list_file = file_map['activity'], file_map['history'], file_map['list']

    st.markdown("\ud83d\udd0d **Files detected:**")
    for key in EXPECTED_FILES:
        st.markdown(f"- {'\u2705' if file_map[key] else '\u274c'} Merchandise {key.capitalize()}")

st.markdown("---")

if st.button("\ud83d\ude80 Run Optimization") and all([activity_file, history_file, list_file]):
    with st.spinner("Processing inventory..."):
        df_activity = pd.read_excel(activity_file)
        df_activity = normalize_columns(df_activity, {
            'qty_expense': 'qty_expensed',
            'wo_qty used': 'wo_qty_used',
            'part': 'partno'
        })

        df_history = pd.read_excel(history_file)
        df_history = normalize_columns(df_history, {
            'part': 'partno', 'part no': 'partno', 'part_no': 'partno',
            'sales': 'qty_sold', 'quantity sold': 'qty_sold', 'tot_qty': 'qty_sold'
        })

        df_list = pd.read_excel(list_file)
        df_list = normalize_columns(df_list, {'part': 'partno', 'part no': 'partno'})
        if 'partno' not in df_list.columns:
            st.error("❌ 'partno' column missing in Merchandise List.")
            st.stop()

        for col in ['qty_sold', 'qty_expensed', 'wo_qty_used']:
            if col not in df_activity.columns:
                df_activity[col] = 0

        df_activity['qty_sold_calc'] = df_activity[['qty_sold', 'qty_expensed', 'wo_qty_used']].sum(axis=1)
        df_activity['max_qty'] = df_activity['qty_sold_calc'] * max_factor
        df_activity['min_qty'] = df_activity['max_qty'] * min_factor
        df_activity['re_order_point'] = df_activity['max_qty'] * reorder_point_factor
        df_activity['re_order_qty'] = df_activity['max_qty'] - df_activity['min_qty']

        for col in ['qty_sold_calc', 'max_qty', 'min_qty', 're_order_point', 're_order_qty']:
            df_activity[col] = df_activity[col].round(0).astype(int)

        if 'partno' not in df_history or 'qty_sold' not in df_history:
            st.error("❌ 'partno' or 'qty_sold' column missing in Merchandise History.")
            st.stop()

        df_history_grouped = df_history.groupby('partno')['qty_sold'].sum().reset_index()
        df_history_grouped['rank'] = df_history_grouped['qty_sold'].rank(ascending=False, method='first')
        df_history_grouped['percentile'] = df_history_grouped['rank'] / len(df_history_grouped)
        df_history_grouped['abc_class'] = df_history_grouped['percentile'].apply(assign_abc)

        df_merge = df_list.merge(
            df_activity[['partno', 'qty_sold_calc', 'max_qty', 'min_qty', 're_order_point', 're_order_qty']],
            on='partno', how='left'
        ).merge(
            df_history_grouped[['partno', 'abc_class']], on='partno', how='left'
        )
        df_merge.rename(columns={'qty_sold_calc': 'qty_sold'}, inplace=True)

        df_merge['sku'] = "NS-" + df_merge['partno'].astype(str).str.replace(" ", "")
        valid = (df_merge['qty_sold'] > 0) & (df_merge['max_qty'] > 0)
        df_merge.loc[valid, 'sku'] = "S-" + df_merge['max_qty'].astype(str) + "-" + df_merge['partno'].astype(str).str.replace(" ", "")
        df_merge['vendor'] = vendor

        drop_cols = ['upc code', 'last purchase date', 'last count date', 'dated added']
        df_merge.drop(columns=[col for col in drop_cols if col in df_merge.columns], inplace=True)

        output = BytesIO()
        df_merge.to_excel(output, index=False)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        col_idx = {cell.value.lower(): cell.column for cell in ws[1]}
        apply_excel_formatting(ws, col_idx)

        output_final = BytesIO()
        wb.save(output_final)
        output_final.seek(0)

        st.success("\u2705 Optimization complete!")
        st.download_button(
            label="\ud83d\udcc5 Download optimized inventory file",
            data=output_final,
            file_name=f"Optimized_Inventory_{vendor.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.subheader("\ud83d\udcca Preview (first 20 rows)")
        st.dataframe(df_merge.head(20))

        st.caption("""
\ud83d\udd34 Red: NS items with past sales  
\ud83d\udfe9 Green: S items with demand  
\u2b1c Grey: No demand or invalid stocking logic  
\ud83d\udfe8 Yellow: Key quantity metrics
""")
else:
    st.info("Please upload all three required files and select a vendor.")
